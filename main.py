def get_element_by_xpath(url):
    if not url:
        return -1
    xpath = '//*[@id="content_main"]/div[1]/div[2]/div[2]/div[1]/div/span/b'
    options = uc.ChromeOptions()
    options.headless = bool(int(config("headless")))  # Запуск в фоновом режиме, если необходимо

    driver = uc.Chrome(options=options)
    driver.get(url)

    try:
        element = driver.find_element(By.XPATH, xpath)
        result = int(element.text.replace(' шт.', ''))
    except NoSuchElementException:
        result = 0
    except Exception:
        traceback.print_exc()
        result = -1
    finally:
        driver.quit()
    return result


def main():
    if not os.path.exists(xlsx_file):
        raise Exception(f"Файл «{xlsx_file}» не найден.")

    wookbook = openpyxl.load_workbook(xlsx_file)
    worksheet = wookbook.active
    for col_i in range(1, worksheet.max_column + 2):
        if worksheet.cell(row=1, column=col_i).value in ['', None]:
            n_col = col_i
            break
    worksheet.cell(row=1, column=n_col).value = datetime.datetime.now(pytz.timezone('Europe/Moscow')).strftime(
        config('time_format'))
    for row_i in range(2, worksheet.max_row + 1):
        url = worksheet.cell(row=row_i, column=3).value
        stock = get_element_by_xpath(url)

        if stock == -1:
            worksheet.cell(row=row_i, column=n_col).value = 'Ошибка'
        else:
            if bool(int(config('console_log'))):
                print(f"{url}: {stock}")
            worksheet.cell(row=row_i, column=n_col).value = stock
        time.sleep(random.randint(1, 10) // 10)
    wookbook.save(xlsx_file)
    print('Скрипт успешно завершил свою работу. ◝(ᵔᵕᵔ)◜')
    print("\nПрограмма завершит свою работу через 120 секунд.")
    time.sleep(120)


if __name__ == '__main__':
    try:
        import datetime
        import os
        import random
        import time
        import traceback
        import warnings

        import undetected_chromedriver as uc
        from selenium.common import NoSuchElementException
        from selenium.webdriver.common.by import By

        warnings.filterwarnings('ignore')
        import openpyxl
        import pytz
        from decouple import config

        xlsx_file = config('xlsx_file')
        main()
    except Exception:
        traceback.print_exc()
        print("\nПрограмма завершит свою работу через 120 секунд.")
        time.sleep(120)
